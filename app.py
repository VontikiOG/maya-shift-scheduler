import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import calendar
import time
import json
import os
import io
import urllib.parse
import streamlit.components.v1 as components
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# הגדרות עמוד נעימות
st.set_page_config(page_title="המשמרות של מאיה", page_icon="🌸", layout="centered")

st.markdown("""
<style>
    .stApp { direction: rtl; }
    p, div, h1, h2, h3, h4, h5, h6, label, span, li, button, input { text-align: right !important; }
    [data-testid="stDataFrame"], [data-testid="stDataEditor"] { direction: rtl; }
    .success-text { color: #2e7d32; font-weight: bold; }
</style>
""", unsafe_allow_html=True)

DRAFT_FILE = "schedule_draft.json"

HOLIDAYS = {
    "02/04/2026": "ערב פסח 🍷",
    "03/04/2026": "פסח 🍷",
    "08/04/2026": "ערב שביעי של פסח",
    "09/04/2026": "שביעי של פסח",
    "21/04/2026": "ערב יום העצמאות",
    "22/04/2026": "יום העצמאות 🇮🇱",
    "21/05/2026": "ערב שבועות",
    "22/05/2026": "שבועות 🌾"
}

# --- פונקציות טיוטה ושמירת מצב ---
def save_draft(data, sandwich_lovers=[]):
    try:
        draft_data = {
            "availability": data,
            "sandwich_lovers": sandwich_lovers
        }
        with open(DRAFT_FILE, 'w', encoding='utf-8') as f:
            json.dump(draft_data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        st.error(f"שגיאה בשמירת הטיוטה: {e}")

def load_draft():
    if os.path.exists(DRAFT_FILE):
        try:
            with open(DRAFT_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if "availability" in data:
                    return data["availability"], data.get("sandwich_lovers", [])
                else:
                    return data, []
        except:
            return {}, []
    return {}, []

# --- פונקציות עזר לחודשים ותאריכים ---
def get_next_month_dates():
    today = datetime.today()
    next_month = today.replace(day=28) + timedelta(days=4) 
    year = next_month.year
    month = next_month.month
    num_days = calendar.monthrange(year, month)[1]
    
    all_dates = []
    weekend_dates = []
    
    for day in range(1, num_days + 1):
        current_date = datetime(year, month, day)
        date_str = current_date.strftime("%d/%m/%Y")
        all_dates.append(date_str)
        if current_date.weekday() in [4, 5]: 
            weekend_dates.append(date_str)
            
    return all_dates, weekend_dates, f"{month:02d}/{year}"

ALL_DATES, WEEKEND_DATES, TARGET_MONTH = get_next_month_dates()

def is_weekend(date_str):
    try:
        return datetime.strptime(date_str, "%d/%m/%Y").weekday() in [4, 5]
    except:
        return False

# --- טעינת נתוני הרופאים ---
@st.cache_data
def load_doctors_data():
    for enc in ['utf-8', 'cp1255', 'iso-8859-8']:
        try:
            df = pd.read_csv("doctors_list.csv", encoding=enc)
            df.columns = df.columns.str.strip()
            if 'מספר טלפון' in df.columns:
                cleaned_phones = df['מספר טלפון'].astype(str).str.replace("'", "").str.replace("-", "").str.strip()
                df['טלפון_נקי'] = cleaned_phones.apply(lambda x: "972" + x[1:] if x.startswith("0") else x)
            return df
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            return pd.DataFrame(columns=["שם הרופא", "מספר טלפון", "טלפון_נקי"])
    return pd.DataFrame(columns=["שם הרופא", "מספר טלפון", "טלפון_נקי"])

# --- אלגוריתם "מד הצדק" המשודרג (v2.2) ---
def generate_fair_schedule(availability_dict, sandwich_lovers):
    all_dates = set()
    for dates in availability_dict.values():
        all_dates.update(dates)
    
    sorted_dates = sorted(list(all_dates), key=lambda x: datetime.strptime(x, "%d/%m/%Y"))
    burnout_scores = {doc: 0 for doc in availability_dict.keys()}
    
    assigned_dates = {doc: set() for doc in availability_dict.keys()}
    avail_counts = {doc: len(dates) for doc, dates in availability_dict.items()}
    
    schedule = []
    bmt_keywords = ["ילנה", "טראסוב", "בהאא", "עתאמנה"]
    unwanted_sandwiches = [] 
    
    def can_work(doc, date_str):
        d = datetime.strptime(date_str, "%d/%m/%Y")
        prev_d = (d - timedelta(days=1)).strftime("%d/%m/%Y")
        next_d = (d + timedelta(days=1)).strftime("%d/%m/%Y")
        return (prev_d not in assigned_dates[doc]) and (next_d not in assigned_dates[doc])

    def creates_sandwich(doc, date_str):
        if doc in sandwich_lovers:
            return False
        d = datetime.strptime(date_str, "%d/%m/%Y")
        for delta in [-3, -2, 2, 3]:
            check_date = (d + timedelta(days=delta)).strftime("%d/%m/%Y")
            if check_date in assigned_dates[doc]:
                return True
        return False

    # --- שלב 1: שיבוץ ראשוני ---
    for date_str in sorted_dates:
        is_we = is_weekend(date_str)
        holiday_note = HOLIDAYS.get(date_str, "")
        
        valid_docs = [doc for doc, dates in availability_dict.items() if date_str in dates and can_work(doc, date_str)]
        
        if not valid_docs:
            schedule.append({
                "תאריך": date_str, "סוג יום": "סופ״ש 🌴" if is_we else "אמצע שבוע ☀️",
                "הערות": holiday_note,
                "המטואונקולוגיה": "חסר רופא 🚨", "השתלות מח עצם": "חסר רופא 🚨"
            })
            continue
            
        bmt_only_docs = [d for d in valid_docs if any(k in d for k in bmt_keywords)]
        regular_docs = [d for d in valid_docs if d not in bmt_only_docs]
        
        bmt_doc = None
        hemato_doc = None
        
        if bmt_only_docs:
            bmt_only_docs.sort(key=lambda d: (
                len(assigned_dates[d]) >= 2,
                creates_sandwich(d, date_str),
                avail_counts[d],
                len(assigned_dates[d]),
                burnout_scores[d]
            ))
            bmt_doc = bmt_only_docs[0]
            
        if regular_docs:
            regular_docs.sort(key=lambda d: (
                len(assigned_dates[d]) >= 2,
                creates_sandwich(d, date_str),
                avail_counts[d],
                len(assigned_dates[d]),
                burnout_scores[d]
            ))
            hemato_doc = regular_docs[0]
            regular_docs.remove(hemato_doc)
            
        if not bmt_doc and regular_docs:
            regular_docs.sort(key=lambda d: (
                len(assigned_dates[d]) >= 2,
                creates_sandwich(d, date_str),
                avail_counts[d],
                len(assigned_dates[d]),
                -burnout_scores[d]
            ))
            bmt_doc = regular_docs[0]
            
        if bmt_doc:
            if creates_sandwich(bmt_doc, date_str) and bmt_doc not in sandwich_lovers:
                unwanted_sandwiches.append((bmt_doc, date_str))
            burnout_scores[bmt_doc] += 1
            assigned_dates[bmt_doc].add(date_str)
            
        if hemato_doc:
            if creates_sandwich(hemato_doc, date_str) and hemato_doc not in sandwich_lovers:
                unwanted_sandwiches.append((hemato_doc, date_str))
            burnout_scores[hemato_doc] += 2
            assigned_dates[hemato_doc].add(date_str)
            
        schedule.append({
            "תאריך": date_str,
            "סוג יום": "סופ״ש 🌴" if is_we else "אמצע שבוע ☀️",
            "הערות": holiday_note,
            "המטואונקולוגיה": hemato_doc if hemato_doc else "חסר רופא 🚨",
            "השתלות מח עצם": bmt_doc if bmt_doc else "חסר רופא 🚨"
        })

    # --- שלב 2: מנגנון "רובין הוד" ---
    for pass_num in [1, 2]:
        for doc in availability_dict.keys():
            while len(assigned_dates[doc]) < 2 and avail_counts[doc] > len(assigned_dates[doc]):
                swapped = False
                for date_str in availability_dict[doc]:
                    if date_str in assigned_dates[doc]: continue
                    if not can_work(doc, date_str): continue
                    
                    if pass_num == 1 and creates_sandwich(doc, date_str):
                        continue 
                    
                    day_dict = next(item for item in schedule if item["תאריך"] == date_str)
                    curr_hemato = day_dict["המטואונקולוגיה"]
                    curr_bmt = day_dict["השתלות מח עצם"]
                    
                    is_bmt_only_doc = any(k in doc for k in bmt_keywords)
                    
                    if not is_bmt_only_doc and curr_hemato != "חסר רופא 🚨" and len(assigned_dates.get(curr_hemato, [])) > 2:
                        day_dict["המטואונקולוגיה"] = doc
                        assigned_dates[curr_hemato].remove(date_str)
                        burnout_scores[curr_hemato] -= 2
                        assigned_dates[doc].add(date_str)
                        burnout_scores[doc] += 2
                        if creates_sandwich(doc, date_str) and doc not in sandwich_lovers:
                            unwanted_sandwiches.append((doc, date_str))
                        swapped = True
                        break
                        
                    if curr_bmt != "חסר רופא 🚨" and len(assigned_dates.get(curr_bmt, [])) > 2:
                        day_dict["השתלות מח עצם"] = doc
                        assigned_dates[curr_bmt].remove(date_str)
                        burnout_scores[curr_bmt] -= 1
                        assigned_dates[doc].add(date_str)
                        burnout_scores[doc] += 1
                        if creates_sandwich(doc, date_str) and doc not in sandwich_lovers:
                            unwanted_sandwiches.append((doc, date_str))
                        swapped = True
                        break
                        
                if not swapped:
                    break

    return pd.DataFrame(schedule), burnout_scores, list(set([doc for doc, d in unwanted_sandwiches]))

# --- פונקציית עיצוב אקסל להורדה (אופציה 1) ---
def create_styled_excel(df, month_name):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet_name = 'סידור חודשי'
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        
        # כיוון הגיליון לימין-שמאל (RTL)
        worksheet.sheet_view.rightToLeft = True
        
        # סגנונות עיצוב
        header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        weekend_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        # עיצוב שורת כותרת
        for cell in worksheet[1]:
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        # עיצוב תאים ושורות סופ"ש
        for row in worksheet.iter_rows(min_row=2, max_col=len(df.columns), max_row=len(df)+1):
            is_weekend = "סופ״ש" in str(row[1].value)
            for cell in row:
                cell.alignment = center_align
                cell.border = thin_border
                if is_weekend:
                    cell.fill = weekend_fill
                    
        # רוחב עמודות אוטומטי
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 25
        worksheet.column_dimensions['D'].width = 25
        worksheet.column_dimensions['E'].width = 25
        
        # הגדרות הדפסה לאקסל: לרוחב והתאמה לעמוד אחד
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 1
        
    return output.getvalue()

# --- פונקציית ייצור HTML להדפסה ישירה (אופציה 2) ---
def get_printable_html(df, month_title):
    html = f"""
    <!DOCTYPE html>
    <html dir="rtl" lang="he">
    <head>
        <meta charset="UTF-8">
        <title>לוח משמרות - {month_title}</title>
        <style>
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #fff; padding: 20px; }}
            .print-btn {{ display: block; margin: 0 auto 20px auto; padding: 12px 24px; font-size: 16px; background-color: #2e7d32; color: white; border: none; border-radius: 8px; cursor: pointer; text-align: center; width: 200px; }}
            .print-btn:hover {{ background-color: #1b5e20; }}
            h2, h3 {{ text-align: center; color: #333; margin: 5px 0; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
            th, td {{ border: 1px solid #ddd; padding: 12px 8px; text-align: center; }}
            th {{ background-color: #D9EAD3; color: #2e7d32; font-weight: bold; font-size: 15px; }}
            tr:nth-child(even) {{ background-color: #fafafa; }}
            .weekend-row {{ background-color: #f0f0f0 !important; font-weight: bold; }}
            .footer {{ text-align: center; margin-top: 30px; font-size: 12px; color: #777; }}
            @media print {{
                .print-btn {{ display: none; }}
                body {{ padding: 0; }}
                table {{ box-shadow: none; border: 2px solid #000; }}
                th, td {{ border: 1px solid #000; color: #000; }}
                th {{ background-color: #e2f0d9 !important; -webkit-print-color-adjust: exact; }}
                .weekend-row {{ background-color: #eaeaea !important; -webkit-print-color-adjust: exact; }}
            }}
        </style>
    </head>
    <body>
        <button class="print-btn" onclick="window.print()">🖨️ הדפס לוח משמרות</button>
        <h2>🏥 לוח תורנויות המטואונקולוגיה והשתלות מח עצם</h2>
        <h3>חודש: {month_title}</h3>
        
        <table>
            <thead>
                <tr>
                    <th>תאריך</th>
                    <th>סוג יום</th>
                    <th>הערות</th>
                    <th>המטואונקולוגיה</th>
                    <th>השתלות מח עצם</th>
                </tr>
            </thead>
            <tbody>
    """
    for _, row in df.iterrows():
        weekend_class = "weekend-row" if "סופ״ש" in str(row['סוג יום']) else ""
        html += f"""
                <tr class="{weekend_class}">
                    <td>{row['תאריך']}</td>
                    <td>{row['סוג יום']}</td>
                    <td>{row.get('הערות', '')}</td>
                    <td>{row['המטואונקולוגיה']}</td>
                    <td>{row['השתלות מח עצם']}</td>
                </tr>
        """
    html += """
            </tbody>
        </table>
        <div class="footer">הופק אוטומטית על ידי מערכת השיבוצים של מאיה 🌸</div>
    </body>
    </html>
    """
    return html

# --- אזור ניהול ויומן שינויים ---
@st.dialog("🔒 כניסת מנהלת")
def admin_login():
    st.markdown("בוקר טוב מאיה! הזיני סיסמה כדי להתחיל לעבוד.")
    pwd = st.text_input("סיסמה", type="password")
    
    correct_password = st.secrets.get("ADMIN_PASSWORD", "PASSWORD_NOT_SET")
    
    if st.button("התחברי 🌸", use_container_width=True):
        if pwd == correct_password and correct_password != "PASSWORD_NOT_SET":
            st.session_state.maya_logged_in = True
            st.rerun()
        elif correct_password == "PASSWORD_NOT_SET":
            st.error("🚨 שגיאת מערכת: מנהל המערכת טרם הגדיר סיסמה בשרת (Secrets).")
        else:
            st.error("סיסמה שגויה, נסי שוב מאיה! 😊")

@st.dialog("📜 יומן שינויים - היסטוריית הפיתוח")
def show_changelog():
    st.markdown("""
    **v3.0.0 | הדפסה מתקדמת ועיצוב 🎨**
    * **אקסל על סטרואידים:** ייצוא קובץ אקסל מעוצב אוטומטית (RTL, כותרות צבעוניות, סימון סופ"שים, מוכן להדפסה בעמוד אחד לרוחב).
    * **הדפסה ישירה:** הוספת תצוגת אינטרנט HTML מרהיבה עם כפתור הדפסה ישיר מתוך האפליקציה עצמה ללא צורך בהורדת אקסל.
    
    **v2.2.0 | מלחמת הסנדוויצ'ים וטבלת סיכום 🥪**
    * מניעת תורנויות צפופות (סנדוויץ'), רשימת חריגים, התראות חכמות, וטבלת סיכום משמרות.

    **v2.1.0 | חוקי ברזל ומינימום משמרות 🛡️**
    * איסור מוחלט על יומיים ברצף, תעדוף רופאים שנתנו מעט תאריכים, מנגנון "רובין הוד" חכם למינימום 2 משמרות.

    **v2.0.0 | חווית משתמש ואבטחה מלאה 🚀**
    * טבלת עריכה ידנית חופשית, וייצור וואטסאפים אוטומטיים לשליחה מהירה.

    **v1.1.0 | טיוטות אוטומטיות ומד הצדק ⚖️**
    * מנגנון טיוטה Auto-Save, ואלגוריתם "נקודות שחיקה".
    """)
    if st.button("סגירה", use_container_width=True):
        st.rerun()

def main():
    if "maya_logged_in" not in st.session_state:
        st.session_state.maya_logged_in = False
    
    if "availability_dict" not in st.session_state:
        avail, lovers = load_draft()
        st.session_state.availability_dict = avail
        st.session_state.sandwich_lovers = lovers

    col1, col2, col3 = st.columns([3, 1, 1])
    with col1:
        st.title("מערכת השיבוצים של מאיה 🌸")
        st.markdown(f"**מכינים את סידור העבודה לחודש: {TARGET_MONTH}**")
    with col2:
        if st.button("מה התחדש?", type="tertiary", use_container_width=True):
            show_changelog()
    with col3:
        if not st.session_state.maya_logged_in:
            if st.button("כניסת מנהלת ⚙️", use_container_width=True):
                admin_login()
        else:
            if st.button("התנתקי 👋", use_container_width=True):
                st.session_state.maya_logged_in = False
                st.rerun()

    st.divider()

    if not st.session_state.maya_logged_in:
        st.info("אנא התחברי למערכת כדי להתחיל לשבץ משמרות.")
        st.caption("*(הטיוטה היומית נשמרת אוטומטית גם אם הדפדפן נסגר)*")
        st.stop()

    df_doctors = load_doctors_data()
    if df_doctors.empty:
        st.warning("⚠️ הקובץ doctors_list.csv חסר. אנא העלי אותו לשרת הגיטהאב.")
        st.stop()
        
    doctor_names = df_doctors['שם הרופא'].tolist()
    
    total_doctors = len(doctor_names)
    fed_doctors_count = len(st.session_state.availability_dict.keys())
    progress_val = min(fed_doctors_count / total_doctors if total_doctors > 0 else 0, 1.0)

    with st.expander("⚙️ הגדרות אילוצים מיוחדים"):
        selected_lovers = st.multiselect(
            "רופאים שמאשרים תורנויות צפופות (סנדוויץ'):",
            doctor_names,
            default=st.session_state.get("sandwich_lovers", [])
        )
        if selected_lovers != st.session_state.get("sandwich_lovers", []):
            st.session_state.sandwich_lovers = selected_lovers
            save_draft(st.session_state.availability_dict, st.session_state.sandwich_lovers)

    st.subheader("👩‍⚕️ הזנת זמינות רופאים")
    st.progress(progress_val, text=f"הוזנו {fed_doctors_count} מתוך {total_doctors} רופאים בקובץ")
    
    selected_doctor = st.selectbox("בחירת רופא להזנה:", ["-- בחרי רופא --"] + doctor_names)
    
    if selected_doctor != "-- בחרי רופא --":
        with st.container(border=True):
            st.markdown(f"#### התאריכים של {selected_doctor}")
            
            is_weekend_only = any(name in selected_doctor for name in ["בהאא", "עתאמנה", "מטקוביץ"])
            is_bmt_only = any(name in selected_doctor for name in ["ילנה", "טראסוב", "בהאא", "עתאמנה"])
            
            if is_weekend_only:
                st.info("💡 שימי לב: רופא זה עובד רק בסופי שבוע. המערכת מציגה לבחירה רק ימי שישי ושבת.")
                options_for_doctor = WEEKEND_DATES
            else:
                options_for_doctor = ALL_DATES
                
            if is_bmt_only:
                st.caption("*(לידיעתך: רופא זה מוגדר למחלקת השתלות מח עצם בלבד)*")
                
            current_selections = st.session_state.availability_dict.get(selected_doctor, [])
            
            def format_date_option(d):
                return f"{d} - {HOLIDAYS[d]}" if d in HOLIDAYS else d
            
            selected_dates = st.multiselect(
                "סמני את התאריכים (אפשר לבחור כמה ביחד):", 
                options_for_doctor,
                default=current_selections,
                format_func=format_date_option
            )
            
            if st.button(f"💾 שמרי תאריכים ל{selected_doctor}", type="primary"):
                if selected_dates:
                    st.session_state.availability_dict[selected_doctor] = selected_dates
                elif selected_doctor in st.session_state.availability_dict:
                    del st.session_state.availability_dict[selected_doctor]
                
                save_draft(st.session_state.availability_dict, st.session_state.sandwich_lovers)
                st.success(f"מעולה! התאריכים של {selected_doctor} נשמרו. ✨")
                time.sleep(1) 
                st.rerun()

    st.divider()
    
    fed_doctors = list(st.session_state.availability_dict.keys())
    
    if fed_doctors:
        with st.expander("📋 רשימת הרופאים שכבר הוזנו החודש"):
            cols = st.columns(3)
            for i, doc in enumerate(fed_doctors):
                num_shifts = len(st.session_state.availability_dict[doc])
                cols[i % 3].markdown(f"✅ {doc} ({num_shifts})")
            
        st.write("")
        if st.button("🪄 צרי סידור עבודה אוטומטי (לפי מד הצדק)", type="primary", use_container_width=True):
            with st.spinner("האלגוריתם מחשב את השיבוץ ההוגן ביותר... ⚖️"):
                schedule_df, burnout_dict, unwanted_sws = generate_fair_schedule(
                    st.session_state.availability_dict, 
                    st.session_state.sandwich_lovers
                )
                st.session_state.final_schedule = schedule_df
                st.session_state.burnout_scores = burnout_dict
                st.session_state.unwanted_sandwiches = unwanted_sws
                st.rerun()
    else:
        st.caption("עדיין לא הוזנו תאריכים לאף רופא.")

    if "final_schedule" in st.session_state:
        st.divider()
        st.subheader("🎉 סידור העבודה מוכן!")
        
        if st.session_state.get("unwanted_sandwiches"):
            docs_str = ", ".join(st.session_state.unwanted_sandwiches)
            st.warning(f"⚠️ שימי לב: נאלצנו לשבץ תורנויות סנדוויץ' לרופאים הבאים כדי לעמוד באילוצים (או להגיע למינימום משמרות): {docs_str}.")
        
        shifts_by_doc = {}
        for idx, row in st.session_state.final_schedule.iterrows():
            date = row['תאריך']
            hemato = row['המטואונקולוגיה']
            bmt = row['השתלות מח עצם']
            
            if hemato and "חסר רופא" not in hemato:
                shifts_by_doc.setdefault(hemato, []).append(f"• {date} (המטואונקולוגיה)")
            if bmt and "חסר רופא" not in bmt:
                shifts_by_doc.setdefault(bmt, []).append(f"• {date} (השתלות מח עצם)")

        summary_data = []
        for doc in fed_doctors:
            summary_data.append({"שם הרופא": doc, "משמרות בפועל": len(shifts_by_doc.get(doc, []))})
        
        summary_df = pd.DataFrame(summary_data).sort_values(by="משמרות בפועל", ascending=False)
        
        with st.container(border=True):
            st.markdown("#### 📊 סיכום כמות משמרות")
            st.dataframe(summary_df, hide_index=True, use_container_width=True)

        st.markdown("אם משהו לא מסתדר לך, **את יכולה ללחוץ על השמות בטבלה ולשנות אותם ידנית** לפני שאת מדפיסה או שולחת.")
        
        edited_df = st.data_editor(st.session_state.final_schedule, use_container_width=True, hide_index=True)
        
        # אזור הדפסה והורדה
        st.divider()
        col_print, col_download = st.columns(2)
        
        with col_print:
            st.subheader("🖨️ הדפסה ישירה")
            st.markdown("להדפסת הלוח בעיצוב נקי ומוכן לתלייה במחלקה:")
            html_content = get_printable_html(edited_df, TARGET_MONTH)
            components.html(html_content, height=120, scrolling=False)
            
        with col_download:
            st.subheader("📥 ייצוא לאקסל")
            st.markdown("להורדת הלוח כקובץ אקסל מעוצב ומוכן להדפסה:")
            excel_data = create_styled_excel(edited_df, TARGET_MONTH)
            st.download_button(
                label="📥 לחצי כאן להורדת סידור מעוצב לאקסל",
                data=excel_data,
                file_name=f"schedule_{TARGET_MONTH.replace('/', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
        
        # חיתוך מחדש במקרה של עריכה ידנית לוואטסאפ
        shifts_by_doc_edited = {}
        for idx, row in edited_df.iterrows():
            date = row['תאריך']
            hemato = row['המטואונקולוגיה']
            bmt = row['השתלות מח עצם']
            
            if hemato and "חסר רופא" not in hemato:
                shifts_by_doc_edited.setdefault(hemato, []).append(f"• {date} (המטואונקולוגיה)")
            if bmt and "חסר רופא" not in bmt:
                shifts_by_doc_edited.setdefault(bmt, []).append(f"• {date} (השתלות מח עצם)")
        
        st.divider()
        st.subheader("📱 שליחת הודעות אישיות לרופאים")
        st.markdown("המערכת הכינה הודעות מרוכזות לכל רופא עם המשמרות שלו:")
        
        phone_dict = dict(zip(df_doctors['שם הרופא'], df_doctors['טלפון_נקי']))
        
        if shifts_by_doc_edited:
            with st.expander("רשימת הודעות מוכנות לשליחה (לחצי לפתיחה)"):
                for doc, shifts in shifts_by_doc_edited.items():
                    phone = phone_dict.get(doc, "")
                    if phone:
                        msg = f"היי ד״ר {doc}, להלן המשמרות שלך לחודש הקרוב:\n" + "\n".join(shifts) + "\n\nתודה רבה, מאיה 🌸"
                        url = f"https://wa.me/{phone}?text={urllib.parse.quote(msg)}"
                        st.link_button(f"💬 שלחי לוואטסאפ של {doc}", url)
                    else:
                        st.warning(f"אין מספר טלפון מעודכן לד״ר {doc}")
                        
    st.write("")
    st.write("")
    with st.expander("⚙️ אפשרויות מתקדמות (איפוס המערכת)"):
        st.warning("לחיצה על הכפתור תמחק את כל התאריכים שהזנת עד כה גם מהגיבוי. הפעולה בלתי הפיכה.")
        if st.button("🗑️ נקי הכל והתחילי מחדש", type="primary"):
            st.session_state.availability_dict = {}
            st.session_state.sandwich_lovers = []
            if "final_schedule" in st.session_state:
                del st.session_state.final_schedule
            if os.path.exists(DRAFT_FILE):
                os.remove(DRAFT_FILE)
            st.success("הכל נוקה בהצלחה! אפשר להתחיל סידור חדש.")
            time.sleep(1)
            st.rerun()

if __name__ == "__main__":
    main()
